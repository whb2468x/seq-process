from __future__ import annotations

import asyncio
import csv
import json
import re
import time
import uuid
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from Bio import Entrez, SeqIO
from Bio.Align import PairwiseAligner
from Bio.Blast import NCBIWWW, NCBIXML
from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

app = FastAPI(title="Sequence Alignment Tool API", version="1.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOADS_DIR = DATA_DIR / "uploads"
RESULTS_DIR = DATA_DIR / "results"
for directory in (UPLOADS_DIR, RESULTS_DIR):
    directory.mkdir(parents=True, exist_ok=True)

SUPPORTED_SEQUENCE_TYPES = {"dna", "rna", "protein"}
NUCLEOTIDE_TYPES = {"dna", "rna"}
PROTEIN_TYPES = {"protein"}
MAX_NCBI_RESULTS = 500

NUCLEOTIDE_ALPHABET = set("ACGTRYSWKMBDHVNU")
PROTEIN_ALPHABET = set("ABCDEFGHIKLMNPQRSTVWXYZUO*")

SUPPORTED_DATABASES: Dict[str, Dict[str, str]] = {
    "nt": {
        "label": "Nucleotide collection (nt)",
        "target_type": "nucleotide",
    },
    "refseq_rna": {
        "label": "RefSeq RNA",
        "target_type": "nucleotide",
    },
    "nr": {
        "label": "Non-redundant protein sequences (nr)",
        "target_type": "protein",
    },
    "swissprot": {
        "label": "Swiss-Prot",
        "target_type": "protein",
    },
    "pdb": {
        "label": "Protein Data Bank (PDB)",
        "target_type": "protein",
    },
}

job_status_store: Dict[str, Dict[str, Any]] = {}


@app.get("/")
async def health_check():
    return {
        "status": "healthy",
        "message": "Sequence Alignment Tool API is running",
        "version": app.version,
    }


@app.get("/databases")
async def get_databases():
    return {
        "databases": {key: value["label"] for key, value in SUPPORTED_DATABASES.items()},
        "metadata": SUPPORTED_DATABASES,
        "max_ncbi_results": MAX_NCBI_RESULTS,
    }


@app.get("/jobs")
async def list_jobs(limit: int = Query(50, ge=1, le=500)):
    jobs = list(job_status_store.values())

    for path in RESULTS_DIR.glob("*.json"):
        try:
            data = json.loads(path.read_text())
        except json.JSONDecodeError:
            continue
        if not any(job.get("job_id") == data.get("job_id") for job in jobs):
            jobs.append(data)

    jobs.sort(key=lambda item: item.get("timestamp") or item.get("created_at") or "", reverse=True)
    return {"jobs": jobs[:limit], "count": min(len(jobs), limit)}


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _result_path(job_id: str, suffix: str) -> Path:
    return RESULTS_DIR / f"{job_id}_{suffix}.json"


def _write_json(path: Path, data: Dict[str, Any]) -> None:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False))


def _parse_fasta_records(content: str) -> List[Any]:
    return list(SeqIO.parse(StringIO(content), "fasta"))


def _extract_sequence_from_text(content: str) -> Tuple[str, Dict[str, Any]]:
    try:
        records = _parse_fasta_records(content)
    except Exception:
        # Parsing failed (e.g. invalid FASTA), treat as raw sequence text
        records = []
    if records:
        first_record = records[0]
        return str(first_record.seq), {
            "record_count": len(records),
            "record_id": first_record.id,
            "input_format": "fasta",
        }

    # Not FASTA format — treat entire content as raw sequence
    return content, {
        "record_count": 1,
        "record_id": None,
        "input_format": "raw",
    }


def _clean_sequence(sequence: str, sequence_type: str) -> str:
    sequence = re.sub(r"\s+", "", sequence).upper()
    sequence = sequence.replace("-", "")

    if not sequence:
        raise HTTPException(status_code=400, detail="Empty sequence provided")

    alphabet = PROTEIN_ALPHABET if sequence_type == "protein" else NUCLEOTIDE_ALPHABET
    invalid = sorted({char for char in sequence if char not in alphabet})
    if invalid:
        raise HTTPException(
            status_code=400,
            detail=f"Sequence contains invalid characters for {sequence_type}: {''.join(invalid[:20])}",
        )

    if sequence_type == "rna":
        return sequence.replace("U", "T")
    return sequence


def _validate_common_params(
    sequence_type: str,
    target_type: str,
    database: Optional[str] = None,
    min_identity: Optional[float] = None,
    max_identity: Optional[float] = None,
    result_count: Optional[int] = None,
) -> None:
    if sequence_type not in SUPPORTED_SEQUENCE_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported sequence_type: {sequence_type}")
    if target_type not in SUPPORTED_SEQUENCE_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported target_type: {target_type}")

    if min_identity is not None and max_identity is not None:
        if not (0 <= min_identity <= 100 and 0 <= max_identity <= 100):
            raise HTTPException(status_code=400, detail="Identity thresholds must be between 0 and 100")
        if min_identity > max_identity:
            raise HTTPException(status_code=400, detail="min_identity cannot be greater than max_identity")

    if result_count is not None and not (1 <= result_count <= MAX_NCBI_RESULTS):
        raise HTTPException(
            status_code=400,
            detail=f"result_count must be between 1 and {MAX_NCBI_RESULTS}",
        )

    if database:
        if database not in SUPPORTED_DATABASES:
            raise HTTPException(status_code=400, detail=f"Database {database} is not supported")
        database_kind = SUPPORTED_DATABASES[database]["target_type"]
        if target_type in NUCLEOTIDE_TYPES and database_kind != "nucleotide":
            raise HTTPException(
                status_code=400,
                detail=f"Database {database} is for protein targets; choose nt/refseq_rna or set target_type=protein",
            )
        if target_type in PROTEIN_TYPES and database_kind != "protein":
            raise HTTPException(
                status_code=400,
                detail=f"Database {database} is for nucleotide targets; choose nr/swissprot/pdb or set target_type=dna/rna",
            )


def choose_blast_program(sequence_type: str, target_type: str) -> str:
    if sequence_type in NUCLEOTIDE_TYPES and target_type in NUCLEOTIDE_TYPES:
        return "blastn"
    if sequence_type in PROTEIN_TYPES and target_type in PROTEIN_TYPES:
        return "blastp"
    if sequence_type in NUCLEOTIDE_TYPES and target_type in PROTEIN_TYPES:
        return "blastx"
    if sequence_type in PROTEIN_TYPES and target_type in NUCLEOTIDE_TYPES:
        return "tblastn"
    raise HTTPException(status_code=400, detail="Unsupported BLAST type combination")


def _alignment_metrics(sequence_a: str, sequence_b: str, sequence_type: str, mode: str = "global") -> Dict[str, Any]:
    clean_a = _clean_sequence(sequence_a, sequence_type)
    clean_b = _clean_sequence(sequence_b, sequence_type)

    aligner = PairwiseAligner()
    aligner.mode = mode
    aligner.match_score = 1.0
    aligner.mismatch_score = 0.0
    aligner.open_gap_score = -1.0
    aligner.extend_gap_score = -0.5

    alignment = aligner.align(clean_a, clean_b)[0]
    coordinates = alignment.coordinates
    aligned_length = 0
    matches = 0

    for index in range(coordinates.shape[1] - 1):
        q0, q1 = int(coordinates[0, index]), int(coordinates[0, index + 1])
        s0, s1 = int(coordinates[1, index]), int(coordinates[1, index + 1])
        q_span = abs(q1 - q0)
        s_span = abs(s1 - s0)
        block_length = max(q_span, s_span)
        aligned_length += block_length

        if q_span and s_span:
            for offset in range(min(q_span, s_span)):
                if clean_a[q0 + offset] == clean_b[s0 + offset]:
                    matches += 1

    identity = (matches / aligned_length) * 100 if aligned_length else 0.0
    return {
        "identity": round(identity, 2),
        "matches": matches,
        "aligned_length": aligned_length,
        "score": round(float(alignment.score), 2),
    }


def calculate_identity(seq1: str, seq2: str) -> float:
    return _alignment_metrics(seq1, seq2, "dna")["identity"]


def _run_ncbi_blast(
    query_sequence: str,
    sequence_type: str,
    target_type: str,
    database: str,
    result_count: int,
    expect_value: float,
    megablast: bool,
    email: Optional[str],
) -> List[Dict[str, Any]]:
    blast_program = choose_blast_program(sequence_type, target_type)
    NCBIWWW.tool = "seq-process"
    if email:
        NCBIWWW.email = email

    kwargs: Dict[str, Any] = {
        "program": blast_program,
        "database": database,
        "sequence": query_sequence,
        "hitlist_size": result_count,
        "expect": expect_value,
        "format_type": "XML",
    }
    if blast_program == "blastn":
        kwargs["megablast"] = megablast

    result_handle = NCBIWWW.qblast(**kwargs)
    results: List[Dict[str, Any]] = []

    try:
        for blast_record in NCBIXML.parse(result_handle):
            for alignment in blast_record.alignments:
                for hsp in alignment.hsps:
                    identity_percent = (hsp.identities / hsp.align_length) * 100 if hsp.align_length else 0.0
                    query_cover = (
                        (abs(hsp.query_end - hsp.query_start) + 1) / blast_record.query_length * 100
                        if blast_record.query_length
                        else 0.0
                    )
                    results.append(
                        {
                            "accession": alignment.accession,
                            "hit_id": alignment.hit_id,
                            "title": alignment.title,
                            "length": alignment.length,
                            "identity": round(identity_percent, 2),
                            "query_cover": round(query_cover, 2),
                            "identities": hsp.identities,
                            "positives": getattr(hsp, "positives", None),
                            "gaps": hsp.gaps,
                            "align_length": hsp.align_length,
                            "score": hsp.score,
                            "bits": hsp.bits,
                            "e_value": hsp.expect,
                            "query_start": hsp.query_start,
                            "query_end": hsp.query_end,
                            "subject_start": hsp.sbjct_start,
                            "subject_end": hsp.sbjct_end,
                            "query_alignment": hsp.query,
                            "subject_alignment": hsp.sbjct,
                            "match_line": hsp.match,
                        }
                    )
    finally:
        result_handle.close()

    return results


def _fetch_sequences_by_accession(
    accessions: List[str],
    entrez_db: str,
    email: str,
    batch_size: int = 100,
) -> Dict[str, str]:
    """Fetch full sequences from NCBI Entrez by accession numbers.

    Returns a dict mapping accession -> cleaned sequence string.
    Sequences that cannot be fetched are silently skipped.
    """
    if not accessions:
        return {}

    Entrez.email = email or "anonymous@example.com"
    Entrez.tool = "seq-process"

    result: Dict[str, str] = {}
    for i in range(0, len(accessions), batch_size):
        batch = accessions[i : i + batch_size]
        try:
            handle = Entrez.efetch(db=entrez_db, id=batch, rettype="fasta", retmode="text")
            records = list(SeqIO.parse(handle, "fasta"))
            handle.close()
            for record in records:
                seq_str = str(record.seq)
                if seq_str:
                    # Store by full record ID (e.g. "JF286587.1")
                    result[record.id] = seq_str
                    # Also store by base accession (strip version suffix if present)
                    # so BLAST accessions (e.g. "JF286587") can match fetched records
                    base_acc = record.id.split(".")[0] if "." in record.id else record.id
                    if base_acc != record.id:
                        result[base_acc] = seq_str
        except Exception:
            pass

        if i + batch_size < len(accessions):
            time.sleep(0.5)

    return result


def _generate_identity_excel(
    identity_matrix: List[List[float]],
    score_matrix: List[List[float]],
    sequence_names: List[str],
    sequence_sources: List[str],
    sequence_accessions: List[str],
    pairwise_details: List[Dict[str, Any]],
) -> bytes:
    """Generate an Excel workbook with identity matrix, pairwise details, and sequence info."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ---- helpers ----
    def _fill(value: float) -> PatternFill:
        if value >= 90:
            return PatternFill(start_color="D73027", end_color="D73027", fill_type="solid")
        if value >= 70:
            return PatternFill(start_color="FC8D59", end_color="FC8D59", fill_type="solid")
        if value >= 50:
            return PatternFill(start_color="FEE08B", end_color="FEE08B", fill_type="solid")
        if value >= 30:
            return PatternFill(start_color="91CF60", end_color="91CF60", fill_type="solid")
        return PatternFill(start_color="1A9850", end_color="1A9850", fill_type="solid")

    def _font(value: float) -> Font:
        return Font(color="FFFFFF", bold=True) if value >= 70 else Font(color="111827", bold=True)

    center = Alignment(horizontal="center", vertical="center")

    # ---- Sheet 1: Identity Matrix ----
    ws1 = wb.active
    ws1.title = "Identity Matrix"

    # Header row
    ws1.cell(row=1, column=1, value="Sequences").font = Font(bold=True)
    for j, name in enumerate(sequence_names):
        cell = ws1.cell(row=1, column=j + 2, value=name)
        cell.font = Font(bold=True)
        cell.alignment = center

    # Data rows
    for i in range(len(sequence_names)):
        row_label = ws1.cell(row=i + 2, column=1, value=sequence_names[i])
        row_label.font = Font(bold=True)
        for j in range(len(sequence_names)):
            value = round(identity_matrix[i][j], 2)
            cell = ws1.cell(row=i + 2, column=j + 2, value=value)
            if i == j:
                # Diagonal: self-comparison
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                cell.font = Font(bold=True)
            else:
                cell.fill = _fill(value)
                cell.font = _font(value)
            cell.alignment = center
            cell.number_format = "0.00"

    # ---- Sheet 2: Score Matrix ----
    ws2 = wb.create_sheet("Score Matrix")
    ws2.cell(row=1, column=1, value="Sequences").font = Font(bold=True)
    for j, name in enumerate(sequence_names):
        ws2.cell(row=1, column=j + 2, value=name).font = Font(bold=True)
    for i in range(len(sequence_names)):
        ws2.cell(row=i + 2, column=1, value=sequence_names[i]).font = Font(bold=True)
        for j in range(len(sequence_names)):
            ws2.cell(row=i + 2, column=j + 2, value=round(score_matrix[i][j], 2)).alignment = center

    # ---- Sheet 3: Pairwise Details ----
    ws3 = wb.create_sheet("Pairwise Details")
    headers3 = ["Sequence A", "Sequence B", "Identity (%)", "Matches", "Aligned Length", "Score"]
    for j, h in enumerate(headers3):
        ws3.cell(row=1, column=j + 1, value=h).font = Font(bold=True)
    for i, detail in enumerate(pairwise_details):
        ws3.cell(row=i + 2, column=1, value=detail.get("sequence_a", ""))
        ws3.cell(row=i + 2, column=2, value=detail.get("sequence_b", ""))
        ws3.cell(row=i + 2, column=3, value=detail.get("identity", 0))
        ws3.cell(row=i + 2, column=4, value=detail.get("matches", 0))
        ws3.cell(row=i + 2, column=5, value=detail.get("aligned_length", 0))
        ws3.cell(row=i + 2, column=6, value=detail.get("score", 0))

    # ---- Sheet 4: Sequence Info ----
    ws4 = wb.create_sheet("Sequence Info")
    info_headers = ["Sequence Name", "Source", "Accession"]
    for j, h in enumerate(info_headers):
        ws4.cell(row=1, column=j + 1, value=h).font = Font(bold=True)
    for i in range(len(sequence_names)):
        ws4.cell(row=i + 2, column=1, value=sequence_names[i])
        ws4.cell(row=i + 2, column=2, value=sequence_sources[i] if i < len(sequence_sources) else "unknown")
        ws4.cell(row=i + 2, column=3, value=sequence_accessions[i] if i < len(sequence_accessions) else "")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _generate_blast_results_excel(
    blast_hits: List[Dict[str, Any]],
    query_name: str = "query",
) -> bytes:
    """Generate an Excel workbook with BLAST hit details (annotation, identity, e-value, etc.)."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BLAST Results"

    headers = [
        "Accession", "Hit ID", "Title", "Length (bp)",
        "Identity (%)", "Query Cover (%)", "E-value", "Bits",
        "Align Length", "Identities", "Positives", "Gaps",
        "Query Start", "Query End", "Subject Start", "Subject End",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for j, h in enumerate(headers):
        cell = ws.cell(row=1, column=j + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, hit in enumerate(blast_hits):
        row_data = [
            hit.get("accession", ""),
            hit.get("hit_id", ""),
            hit.get("title", ""),
            hit.get("length", ""),
            round(hit.get("identity", 0), 2),
            round(hit.get("query_cover", 0), 2),
            f"{hit.get('e_value', 0):.2e}",
            hit.get("bits", ""),
            hit.get("align_length", ""),
            hit.get("identities", ""),
            hit.get("positives", ""),
            hit.get("gaps", ""),
            hit.get("query_start", ""),
            hit.get("query_end", ""),
            hit.get("subject_start", ""),
            hit.get("subject_end", ""),
        ]
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=val).alignment = center

    # Adjust column widths
    col_widths = [16, 18, 50, 12, 14, 16, 16, 10, 14, 12, 12, 8, 14, 12, 16, 14]
    for j, w in enumerate(col_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _generate_sequences_fasta_bytes(
    sequence_names: List[str],
    sequences: List[str],
) -> bytes:
    """Generate a FASTA-formatted string with all sequences."""
    lines: List[str] = []
    for name, seq in zip(sequence_names, sequences):
        lines.append(f">{name}")
        # Wrap sequence at 70 chars per line
        for i in range(0, len(seq), 70):
            lines.append(seq[i : i + 70])
        lines.append("")
    return ("\n".join(lines)).encode("utf-8")


def _generate_heatmap_image(
    identity_matrix: List[List[float]],
    sequence_names: List[str],
) -> bytes:
    """Generate a heatmap PNG image from the identity matrix."""
    if not HAS_MATPLOTLIB:
        raise RuntimeError("matplotlib is not installed. Run: pip install matplotlib")

    import numpy as np
    n = len(sequence_names)
    data = np.array(identity_matrix)

    # Figure size proportional to matrix size
    fig_size = max(6, min(20, n * 0.5))
    fig, ax = plt.subplots(figsize=(fig_size + 2, fig_size))

    # Custom colormap: green (low) -> yellow -> orange -> red (high)
    colors_list = ["#1a9850", "#91cf60", "#fee08b", "#fc8d59", "#d73027"]
    cmap = mcolors.LinearSegmentedColormap.from_list("identity_cmap", colors_list)

    im = ax.imshow(data, cmap=cmap, vmin=0, vmax=100, aspect="auto")

    # Colorbar
    cbar = fig.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label("Identity (%)", fontsize=12)

    # Labels
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    # Truncate long names
    short_names = [name if len(name) <= 18 else name[:15] + "..." for name in sequence_names]
    ax.set_xticklabels(short_names, rotation=45, ha="right", fontsize=8)
    ax.set_yticklabels(short_names, fontsize=8)

    # Add text annotations for small matrices
    if n <= 30:
        for i in range(n):
            for j in range(n):
                text_color = "white" if data[i, j] >= 70 else "black"
                ax.text(j, i, f"{data[i, j]:.1f}", ha="center", va="center",
                        fontsize=7, color=text_color)

    ax.set_title("Pairwise Sequence Identity Heatmap", fontsize=14, fontweight="bold")
    plt.tight_layout()

    output = BytesIO()
    fig.savefig(output, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    output.seek(0)
    return output.getvalue()


def _generate_filtered_sequences_log(
    filtered_entries: List[Dict[str, Any]],
) -> str:
    """Generate a text log of filtered sequences with reasons."""
    lines = [
        f"Filtered Sequences Log - {_now()}",
        f"Total filtered: {len(filtered_entries)}",
        "=" * 60,
        "",
    ]
    for i, entry in enumerate(filtered_entries, 1):
        lines.append(f"{i}. {entry.get('accession', 'unknown')}")
        lines.append(f"   Reason: {entry.get('reason', 'unknown')}")
        if "detail" in entry:
            lines.append(f"   Detail: {entry['detail']}")
        lines.append("")
    return "\n".join(lines)


def _run_combined_pairwise_alignment(
    content_str: str,
    sequence_type: str,
    target_type: str,
    database: str,
    alignment_mode: str,
    min_identity: float,
    max_identity: float,
    result_count: int,
    expect_value: float,
    megablast: bool,
    email: Optional[str],
    max_total_sequences: int,
) -> Dict[str, Any]:
    """Core combined pipeline: BLAST user sequences, fetch hits, and run pairwise.

    This is a synchronous function designed to run inside asyncio.to_thread.

    Returns a dict with:
      - combined_seqs, combined_names, combined_sources, combined_accessions (raw data)
      - identity_matrix, score_matrix, heatmap_data (for file generation, NOT in JSON)
      - pairwise_details (for JSON + CSV)
      - blast_hits_all (all BLAST hit metadata for Excel generation)
      - filtered_sequences (list of filtered accessions with reasons)
    """
    filtered_entries: List[Dict[str, Any]] = []

    # 1. Parse user sequences
    records = _parse_fasta_records(content_str)
    if len(records) < 1:
        raise HTTPException(status_code=400, detail="At least 1 FASTA sequence is required")

    user_sequences: List[str] = []
    user_names: List[str] = []
    used_names: Dict[str, int] = {}
    for record in records:
        clean_seq = _clean_sequence(str(record.seq), sequence_type)
        if clean_seq:
            user_sequences.append(clean_seq)
            user_names.append(_unique_name(record.id, used_names))
    if not user_sequences:
        raise HTTPException(status_code=400, detail="No valid sequences found in file")

    n_user = len(user_sequences)

    # 2. Determine Entrez database type
    entrez_db = "nucleotide" if target_type in NUCLEOTIDE_TYPES else "protein"

    # 3. Run BLAST on each user sequence (up to 5), collect unique accessions AND hit details
    max_blast_queries = min(n_user, 5)
    results_per_query = max(1, min(result_count, 50) // max_blast_queries)
    all_accessions: set = set()
    blast_hit_count = 0
    blast_hits_all: List[Dict[str, Any]] = []

    for idx in range(max_blast_queries):
        try:
            raw_results = _run_ncbi_blast(
                user_sequences[idx],
                sequence_type,
                target_type,
                database,
                results_per_query,
                expect_value,
                megablast,
                email,
            )
            for result in raw_results:
                acc = result.get("accession", "").strip()
                if not acc:
                    continue
                identity = result["identity"]
                if not (min_identity <= identity <= max_identity):
                    filtered_entries.append({
                        "accession": acc,
                        "reason": "identity_filter",
                        "detail": f"Identity {identity}% outside range [{min_identity}%, {max_identity}%]",
                        "title": result.get("title", ""),
                    })
                    continue
                if acc not in all_accessions:
                    all_accessions.add(acc)
                    blast_hits_all.append(dict(result))
            blast_hit_count += len(raw_results)
        except Exception as exc:
            filtered_entries.append({
                "accession": f"query_{idx}",
                "reason": "blast_error",
                "detail": str(exc),
                "title": user_names[idx] if idx < len(user_names) else "unknown",
            })

    # 4. Fetch full sequences from NCBI via Entrez
    accessions_list = list(all_accessions)
    max_blast_seqs = max_total_sequences - n_user
    truncated_count = 0
    if len(accessions_list) > max_blast_seqs:
        truncated = accessions_list[max_blast_seqs:]
        for acc in truncated:
            filtered_entries.append({
                "accession": acc,
                "reason": "count_limit",
                "detail": f"Exceeded max_total_sequences ({max_total_sequences}); "
                          f"only first {max_blast_seqs} NCBI sequences kept",
            })
        truncated_count = len(truncated)
        accessions_list = accessions_list[:max_blast_seqs]

    fetched_sequences: Dict[str, str] = {}
    fetch_failed: List[str] = []
    if accessions_list:
        try:
            fetched_sequences = _fetch_sequences_by_accession(
                accessions_list, entrez_db, email or ""
            )
        except Exception as exc:
            filtered_entries.append({
                "accession": "batch",
                "reason": "fetch_error",
                "detail": str(exc),
            })
        # Track fetch failures
        for acc in accessions_list:
            if acc not in fetched_sequences:
                fetch_failed.append(acc)

    for acc in fetch_failed:
        filtered_entries.append({
            "accession": acc,
            "reason": "fetch_failed",
            "detail": "Could not fetch full sequence from NCBI Entrez",
        })

    # 5. Combine: user sequences + fetched NCBI sequences
    combined_seqs = list(user_sequences)
    combined_names = list(user_names)
    combined_sources = ["user"] * n_user
    combined_accessions = [""] * n_user

    # Iterate over original BLAST accessions to avoid duplicates
    # (fetched_sequences may contain both versioned and base keys)
    seen_base_accessions: set = set()
    for base_acc in accessions_list:
        seq = fetched_sequences.get(base_acc)
        if seq is None:
            continue  # truly failed to fetch

        # Skip if the base accession already processed (avoid versioned duplicates)
        if base_acc in seen_base_accessions:
            continue
        seen_base_accessions.add(base_acc)

        # Skip if the sequence is too long (>50k chars) for practical pairwise
        if len(seq) > 50000:
            filtered_entries.append({
                "accession": base_acc,
                "reason": "sequence_too_long",
                "detail": f"Sequence length {len(seq)} exceeds 50,000 bp limit",
            })
            continue
        # Use base accession as display name (cleaner than versioned form)
        display_acc = base_acc
        if display_acc in used_names:
            display_acc = _unique_name(display_acc, used_names)
            filtered_entries.append({
                "accession": base_acc,
                "reason": "duplicate_name",
                "detail": f"Renamed to {display_acc}",
            })
        else:
            used_names[display_acc] = 1
        combined_seqs.append(seq)
        combined_names.append(display_acc)
        combined_sources.append("ncbi")
        combined_accessions.append(display_acc)

    # If we fetched nothing, at least run pairwise on user's own sequences
    if len(combined_seqs) < 2:
        raise HTTPException(
            status_code=400,
            detail="Not enough sequences for pairwise comparison. "
            "Try lowering identity filters or increasing result count.",
        )

    # 6. Run pairwise alignment on combined set
    n = len(combined_seqs)
    identity_matrix = [[100.0 for _ in range(n)] for _ in range(n)]
    score_matrix = [[0.0 for _ in range(n)] for _ in range(n)]
    heatmap_data: List[Dict[str, Any]] = []
    pairwise_details: List[Dict[str, Any]] = []

    for i in range(n):
        for j in range(i, n):
            if i == j:
                metrics = {
                    "identity": 100.0,
                    "matches": len(combined_seqs[i]),
                    "aligned_length": len(combined_seqs[i]),
                    "score": float(len(combined_seqs[i])),
                }
            else:
                metrics = _alignment_metrics(
                    combined_seqs[i], combined_seqs[j], sequence_type, alignment_mode
                )

            identity_matrix[i][j] = identity_matrix[j][i] = metrics["identity"]
            score_matrix[i][j] = score_matrix[j][i] = metrics["score"]

            if i < j:
                pairwise_details.append(
                    {
                        "sequence_a": combined_names[i],
                        "sequence_b": combined_names[j],
                        "identity": metrics["identity"],
                        "matches": metrics["matches"],
                        "aligned_length": metrics["aligned_length"],
                        "score": metrics["score"],
                        "source_a": combined_sources[i],
                        "source_b": combined_sources[j],
                    }
                )

    for i in range(n):
        for j in range(n):
            heatmap_data.append(
                {"x": combined_names[j], "y": combined_names[i], "value": identity_matrix[i][j]}
            )

    return {
        # Raw sequence data (for file generation, NOT included in JSON response)
        "_combined_seqs": combined_seqs,
        "_identity_matrix": identity_matrix,
        "_score_matrix": score_matrix,
        "_heatmap_data": heatmap_data,
        "_blast_hits_all": blast_hits_all,
        # Lightweight JSON fields
        "combined_names": combined_names,
        "combined_sources": combined_sources,
        "combined_accessions": combined_accessions,
        "n_total": n,
        "n_user": n_user,
        "n_ncbi_fetched": len([a for a in accessions_list if a in fetched_sequences]),
        "n_blast_hits": blast_hit_count,
        "n_filtered": len(filtered_entries),
        "n_truncated": truncated_count,
        "n_fetch_failed": len(fetch_failed),
        "pairwise_details": pairwise_details,
        "filtered_sequences": filtered_entries,
    }


async def run_combined_pairwise_analysis(
    job_id: str,
    content_str: str,
    sequence_type: str,
    target_type: str,
    database: str,
    alignment_mode: str,
    min_identity: float,
    max_identity: float,
    result_count: int,
    expect_value: float,
    megablast: bool,
    email: Optional[str],
    max_total_sequences: int,
):
    """Async wrapper for the combined BLAST+pairwise pipeline with progress tracking."""
    blast_program = choose_blast_program(sequence_type, target_type)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _update_progress(progress: str, message: str) -> None:
        if job_id in job_status_store:
            job_status_store[job_id]["progress"] = progress
            job_status_store[job_id]["message"] = message

    job_status_store[job_id] = {
        "job_id": job_id,
        "status": "processing",
        "progress": "starting",
        "message": f"Starting combined analysis: BLAST ({blast_program}) → Fetch → Pairwise",
        "program": blast_program,
        "database": database,
        "analysis_type": "pairwise_blast",
        "has_excel": False,
        "has_fasta": False,
        "has_blast_excel": False,
        "has_heatmap": False,
        "has_filter_log": False,
        "created_at": _now(),
    }

    try:
        # ----- Step 1: Run the combined pipeline (BLAST + Fetch + Pairwise) -----
        _update_progress("running_blast", f"Running NCBI BLAST ({blast_program}) against {database}...")
        raw = await asyncio.to_thread(
            _run_combined_pairwise_alignment,
            content_str,
            sequence_type,
            target_type,
            database,
            alignment_mode,
            min_identity,
            max_identity,
            result_count,
            expect_value,
            megablast,
            email,
            max_total_sequences,
        )
        _update_progress("generating_files", "Generating result files (FASTA, Excel, Heatmap)...")

        # Extract raw data
        combined_seqs = raw.pop("_combined_seqs", [])
        identity_matrix = raw.pop("_identity_matrix", [])
        score_matrix = raw.pop("_score_matrix", [])
        heatmap_data_list = raw.pop("_heatmap_data", [])
        blast_hits_all = raw.pop("_blast_hits_all", [])

        # Build lightweight JSON result
        result_data = {
            "job_id": job_id,
            "status": "completed",
            "analysis_type": "pairwise_blast",
            "message": f"Combined analysis completed: "
                       f"{raw['n_user']} user + {raw['n_ncbi_fetched']} NCBI "
                       f"→ {raw['n_total']} total sequences ({raw['n_filtered']} filtered)",
            "sequence_type": sequence_type,
            "target_type": target_type,
            "database": database,
            "alignment_mode": alignment_mode,
            "blast_program": blast_program,
            "user_sequence_count": raw["n_user"],
            "ncbi_sequence_count": raw["n_ncbi_fetched"],
            "blast_hit_count": raw["n_blast_hits"],
            "total_sequences": raw["n_total"],
            "filtered_count": raw["n_filtered"],
            "filtered_sequences": raw["filtered_sequences"],
            "sequence_names": raw["combined_names"],
            "sequence_sources": raw["combined_sources"],
            "sequence_accessions": raw["combined_accessions"],
            "sequence_stats": [
                _sequence_stats(raw["combined_names"][i], combined_seqs[i], sequence_type)
                for i in range(raw["n_total"])
            ],
            "pairwise_details": raw["pairwise_details"],
            "has_excel": False,
            "has_fasta": False,
            "has_blast_excel": False,
            "has_heatmap": False,
            "has_filter_log": False,
            "timestamp": _now(),
        }

        # ----- Step 2: Generate FASTA file (all sequences) -----
        _update_progress("generating_files", "Writing FASTA file...")
        try:
            fasta_bytes = await asyncio.to_thread(
                _generate_sequences_fasta_bytes,
                raw["combined_names"],
                combined_seqs,
            )
            fasta_path = _result_path(job_id, "combined")
            fasta_path.with_suffix(".fasta").write_bytes(fasta_bytes)
            result_data["has_fasta"] = True
        except Exception as exc:
            result_data["fasta_error"] = str(exc)

        # ----- Step 3: Generate BLAST results Excel -----
        if HAS_OPENPYXL and blast_hits_all:
            _update_progress("generating_files", "Writing BLAST results Excel...")
            try:
                blast_excel_bytes = await asyncio.to_thread(
                    _generate_blast_results_excel,
                    blast_hits_all,
                    raw["combined_names"][0] if raw["combined_names"] else "query",
                )
                blast_excel_path = _result_path(job_id, "combined_blast")
                blast_excel_path.with_suffix(".xlsx").write_bytes(blast_excel_bytes)
                result_data["has_blast_excel"] = True
            except Exception as exc:
                result_data["blast_excel_error"] = str(exc)

        # ----- Step 4: Generate Identity Matrix Excel -----
        if HAS_OPENPYXL:
            _update_progress("generating_files", "Writing identity matrix Excel...")
            try:
                excel_bytes = await asyncio.to_thread(
                    _generate_identity_excel,
                    identity_matrix,
                    score_matrix,
                    raw["combined_names"],
                    raw["combined_sources"],
                    raw["combined_accessions"],
                    raw["pairwise_details"],
                )
                excel_path = _result_path(job_id, "combined")
                excel_path.with_suffix(".xlsx").write_bytes(excel_bytes)
                result_data["has_excel"] = True
            except Exception as exc:
                result_data["excel_error"] = str(exc)

        # ----- Step 5: Generate Heatmap PNG -----
        if HAS_MATPLOTLIB and raw["n_total"] >= 2:
            _update_progress("generating_files", "Rendering heatmap image...")
            try:
                heatmap_bytes = await asyncio.to_thread(
                    _generate_heatmap_image,
                    identity_matrix,
                    raw["combined_names"],
                )
                heatmap_path = _result_path(job_id, "combined")
                heatmap_path.with_suffix(".png").write_bytes(heatmap_bytes)
                result_data["has_heatmap"] = True
            except Exception as exc:
                result_data["heatmap_error"] = str(exc)

        # ----- Step 6: Generate filter log -----
        if raw["filtered_sequences"]:
            try:
                filter_log = _generate_filtered_sequences_log(raw["filtered_sequences"])
                filter_path = _result_path(job_id, "combined_filter")
                filter_path.with_suffix(".txt").write_text(filter_log)
                result_data["has_filter_log"] = True
            except Exception as exc:
                result_data["filter_log_error"] = str(exc)

        _update_progress("completed", result_data["message"])
        _write_json(_result_path(job_id, "pairwise_blast_results"), result_data)
        result_data["progress"] = "completed"
        job_status_store[job_id] = result_data

    except HTTPException:
        error_data = {
            "job_id": job_id,
            "status": "failed",
            "analysis_type": "pairwise_blast",
            "program": blast_program,
            "database": database,
            "has_excel": False,
            "has_fasta": False,
            "has_blast_excel": False,
            "has_heatmap": False,
            "has_filter_log": False,
            "message": f"Combined analysis rejected: invalid input parameters",
            "timestamp": _now(),
        }
        _write_json(_result_path(job_id, "pairwise_blast_error"), error_data)
        job_status_store[job_id] = error_data
    except Exception as exc:
        error_data = {
            "job_id": job_id,
            "status": "failed",
            "analysis_type": "pairwise_blast",
            "program": blast_program,
            "database": database,
            "has_excel": False,
            "has_fasta": False,
            "has_blast_excel": False,
            "has_heatmap": False,
            "has_filter_log": False,
            "message": f"Combined analysis failed: {exc}",
            "error": str(exc),
            "timestamp": _now(),
        }
        _write_json(_result_path(job_id, "pairwise_blast_error"), error_data)
        job_status_store[job_id] = error_data


async def run_blast_analysis(
    job_id: str,
    query_sequence: str,
    sequence_type: str,
    target_type: str,
    database: str,
    min_identity: float,
    max_identity: float,
    result_count: int,
    expect_value: float = 10.0,
    megablast: bool = True,
    email: Optional[str] = None,
    input_metadata: Optional[Dict[str, Any]] = None,
):
    blast_program = choose_blast_program(sequence_type, target_type)
    job_status_store[job_id] = {
        "job_id": job_id,
        "status": "processing",
        "message": f"Calling NCBI BLAST with {blast_program}",
        "program": blast_program,
        "database": database,
        "analysis_type": "blast",
        "has_blast_excel": False,
        "created_at": _now(),
    }

    try:
        raw_results = await asyncio.to_thread(
            _run_ncbi_blast,
            query_sequence,
            sequence_type,
            target_type,
            database,
            result_count,
            expect_value,
            megablast,
            email,
        )

        filtered_results = [
            result for result in raw_results if min_identity <= result["identity"] <= max_identity
        ][:result_count]

        result_data = {
            "job_id": job_id,
            "status": "completed",
            "message": f"BLAST analysis completed. Found {len(filtered_results)} matching HSPs.",
            "analysis_type": "blast",
            "program": blast_program,
            "database": database,
            "sequence_type": sequence_type,
            "target_type": target_type,
            "query_length": len(query_sequence),
            "filters": {
                "min_identity": min_identity,
                "max_identity": max_identity,
                "result_count": result_count,
                "expect_value": expect_value,
            },
            "input": input_metadata or {},
            "results": filtered_results,
            "result_count": len(filtered_results),
            "has_blast_excel": False,
            "timestamp": _now(),
        }

        # Generate BLAST results Excel
        if HAS_OPENPYXL and filtered_results:
            try:
                excel_bytes = await asyncio.to_thread(
                    _generate_blast_results_excel,
                    filtered_results,
                    input_metadata.get("record_id", "query") if input_metadata else "query",
                )
                excel_path = _result_path(job_id, "blast")
                excel_path.with_suffix(".xlsx").write_bytes(excel_bytes)
                result_data["has_blast_excel"] = True
            except Exception:
                pass

        _write_json(_result_path(job_id, "blast_results"), result_data)
        job_status_store[job_id] = result_data

    except HTTPException:
        error_data = {
            "job_id": job_id,
            "status": "failed",
            "analysis_type": "blast",
            "program": blast_program,
            "database": database,
            "has_blast_excel": False,
            "message": "BLAST analysis rejected: invalid input parameters",
            "timestamp": _now(),
        }
        _write_json(_result_path(job_id, "error"), error_data)
        job_status_store[job_id] = error_data
    except Exception as exc:
        error_data = {
            "job_id": job_id,
            "status": "failed",
            "analysis_type": "blast",
            "program": blast_program,
            "database": database,
            "has_blast_excel": False,
            "message": f"BLAST analysis failed: {exc}",
            "error": str(exc),
            "timestamp": _now(),
        }
        _write_json(_result_path(job_id, "error"), error_data)
        job_status_store[job_id] = error_data


@app.post("/submit-sequence")
async def submit_sequence(
    sequence: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    sequence_type: str = Form("dna"),
    target_type: str = Form("dna"),
    database: str = Form("nt"),
    min_identity: float = Form(30.0),
    max_identity: float = Form(100.0),
    result_count: int = Form(100),
    expect_value: float = Form(10.0),
    megablast: bool = Form(True),
    email: Optional[str] = Form(None),
):
    _validate_common_params(
        sequence_type,
        target_type,
        database=database,
        min_identity=min_identity,
        max_identity=max_identity,
        result_count=result_count,
    )
    if expect_value <= 0:
        raise HTTPException(status_code=400, detail="expect_value must be greater than 0")
    if not sequence and not file:
        raise HTTPException(status_code=400, detail="Either sequence or file must be provided")

    input_metadata: Dict[str, Any] = {"source": "text"}
    if sequence:
        raw_sequence, parsed_metadata = _extract_sequence_from_text(sequence.strip())
        input_metadata.update(parsed_metadata)
    else:
        assert file is not None
        content = await file.read()
        try:
            content_str = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail="Uploaded file must be UTF-8 text") from exc

        job_upload_id = str(uuid.uuid4())
        safe_name = Path(file.filename or "input.fasta").name
        upload_path = UPLOADS_DIR / f"{job_upload_id}_{safe_name}"
        upload_path.write_bytes(content)

        raw_sequence, parsed_metadata = _extract_sequence_from_text(content_str)
        input_metadata.update(parsed_metadata)
        input_metadata.update({"source": "file", "filename": safe_name, "saved_path": str(upload_path)})

    try:
        query_sequence = _clean_sequence(raw_sequence, sequence_type)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to clean sequence: {exc}",
        ) from exc
    job_id = str(uuid.uuid4())

    asyncio.create_task(
        run_blast_analysis(
            job_id,
            query_sequence,
            sequence_type,
            target_type,
            database,
            min_identity,
            max_identity,
            result_count,
            expect_value,
            megablast,
            email,
            input_metadata,
        )
    )

    return {
        "job_id": job_id,
        "status": "submitted",
        "message": "BLAST analysis started",
        "program": choose_blast_program(sequence_type, target_type),
        "query_length": len(query_sequence),
    }


def _load_job_result(job_id: str) -> Dict[str, Any]:
    if job_id in job_status_store:
        return job_status_store[job_id]

    for suffix in (
        "blast_results",
        "pairwise_results",
        "pairwise_blast_results",
        "pairwise_blast_error",
        "error",
    ):
        path = _result_path(job_id, suffix)
        if path.exists():
            data = json.loads(path.read_text())
            job_status_store[job_id] = data
            return data

    raise HTTPException(status_code=404, detail="Job not found")


@app.get("/job-status/{job_id}")
async def get_job_status(job_id: str):
    return _load_job_result(job_id)


def _rows_to_csv(rows: Iterable[Dict[str, Any]]) -> StringIO:
    output = StringIO()
    rows = list(rows)
    if not rows:
        output.write("")
        output.seek(0)
        return output

    fieldnames = list(rows[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return output


@app.get("/results/{job_id}")
async def download_result(job_id: str, format: str = Query("json", pattern="^(json|csv)$")):
    result = _load_job_result(job_id)
    filename = f"{job_id}.{format}"

    if format == "json":
        return JSONResponse(
            result,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if result.get("analysis_type") in ("pairwise", "pairwise_blast"):
        rows = result.get("pairwise_details", [])
    else:
        rows = result.get("results", [])

    output = _rows_to_csv(rows)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _unique_name(name: str, used_names: Dict[str, int]) -> str:
    base_name = name or f"seq_{len(used_names) + 1}"
    if base_name not in used_names:
        used_names[base_name] = 1
        return base_name

    used_names[base_name] += 1
    return f"{base_name}_{used_names[base_name]}"


def _sequence_stats(name: str, sequence: str, sequence_type: str) -> Dict[str, Any]:
    stats = {"name": name, "length": len(sequence)}
    if sequence_type in NUCLEOTIDE_TYPES:
        gc_count = sequence.count("G") + sequence.count("C")
        stats["gc_percent"] = round((gc_count / len(sequence)) * 100, 2) if sequence else 0.0
    return stats


def _run_pairwise_alignment(
    content_str: str,
    sequence_type: str,
    alignment_mode: str,
    max_sequences: int,
) -> Dict[str, Any]:
    records = _parse_fasta_records(content_str)
    if len(records) < 2:
        raise HTTPException(status_code=400, detail="At least 2 FASTA sequences are required")
    if len(records) > max_sequences:
        raise HTTPException(status_code=400, detail=f"Too many sequences; limit is {max_sequences}")

    sequences: List[str] = []
    names: List[str] = []
    used_names: Dict[str, int] = {}

    for record in records:
        clean_sequence = _clean_sequence(str(record.seq), sequence_type)
        sequences.append(clean_sequence)
        names.append(_unique_name(record.id, used_names))

    n = len(sequences)
    identity_matrix = [[100.0 for _ in range(n)] for _ in range(n)]
    score_matrix = [[0.0 for _ in range(n)] for _ in range(n)]
    heatmap_data: List[Dict[str, Any]] = []
    pairwise_details: List[Dict[str, Any]] = []

    for i in range(n):
        for j in range(i, n):
            if i == j:
                metrics = {
                    "identity": 100.0,
                    "matches": len(sequences[i]),
                    "aligned_length": len(sequences[i]),
                    "score": float(len(sequences[i])),
                }
            else:
                metrics = _alignment_metrics(sequences[i], sequences[j], sequence_type, alignment_mode)

            identity_matrix[i][j] = identity_matrix[j][i] = metrics["identity"]
            score_matrix[i][j] = score_matrix[j][i] = metrics["score"]

            if i < j:
                pairwise_details.append(
                    {
                        "sequence_a": names[i],
                        "sequence_b": names[j],
                        "identity": metrics["identity"],
                        "matches": metrics["matches"],
                        "aligned_length": metrics["aligned_length"],
                        "score": metrics["score"],
                    }
                )

    for i in range(n):
        for j in range(n):
            heatmap_data.append({"x": names[j], "y": names[i], "value": identity_matrix[i][j]})

    job_id = str(uuid.uuid4())
    return {
        "job_id": job_id,
        "status": "completed",
        "analysis_type": "pairwise",
        "message": f"Pairwise alignment completed for {n} sequences",
        "sequence_type": sequence_type,
        "alignment_mode": alignment_mode,
        "sequence_names": names,
        "sequence_stats": [_sequence_stats(names[i], sequences[i], sequence_type) for i in range(n)],
        "identity_matrix": identity_matrix,
        "score_matrix": score_matrix,
        "heatmap_data": heatmap_data,
        "pairwise_details": pairwise_details,
        "timestamp": _now(),
    }


@app.post("/pairwise-alignment")
async def pairwise_alignment(
    sequence: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    sequence_type: str = Form("dna"),
    alignment_mode: str = Form("global"),
    max_sequences: int = Form(200),
):
    _validate_common_params(sequence_type, sequence_type)
    if alignment_mode not in {"global", "local"}:
        raise HTTPException(status_code=400, detail="alignment_mode must be global or local")
    if not (2 <= max_sequences <= 1000):
        raise HTTPException(status_code=400, detail="max_sequences must be between 2 and 1000")
    if not sequence and not file:
        raise HTTPException(status_code=400, detail="Either sequence text (FASTA) or file must be provided")

    input_info: Dict[str, Any] = {}
    if sequence:
        content_str = sequence.strip()
        input_info["source"] = "text"
    else:
        assert file is not None
        content = await file.read()
        try:
            content_str = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail="Uploaded file must be UTF-8 text") from exc
        safe_name = Path(file.filename or "pairwise.fasta").name
        upload_path = UPLOADS_DIR / f"{uuid.uuid4()}_{safe_name}"
        upload_path.write_bytes(content)
        input_info["source"] = "file"
        input_info["filename"] = safe_name
        input_info["saved_path"] = str(upload_path)

    result = await asyncio.to_thread(
        _run_pairwise_alignment,
        content_str,
        sequence_type,
        alignment_mode,
        max_sequences,
    )
    result["input"] = input_info

    # Generate Excel if available
    if HAS_OPENPYXL:
        try:
            sources = ["user"] * len(result["sequence_names"])
            accessions = [""] * len(result["sequence_names"])
            excel_bytes = await asyncio.to_thread(
                _generate_identity_excel,
                result["identity_matrix"],
                result["score_matrix"],
                result["sequence_names"],
                sources,
                accessions,
                result["pairwise_details"],
            )
            excel_path = _result_path(result["job_id"], "pairwise")
            excel_path.with_suffix(".xlsx").write_bytes(excel_bytes)
            result["has_excel"] = True
        except Exception:
            result["has_excel"] = False

    # Generate heatmap if available
    if HAS_MATPLOTLIB and len(result["sequence_names"]) >= 2:
        try:
            heatmap_bytes = await asyncio.to_thread(
                _generate_heatmap_image,
                result["identity_matrix"],
                result["sequence_names"],
            )
            heatmap_path = _result_path(result["job_id"], "pairwise")
            heatmap_path.with_suffix(".png").write_bytes(heatmap_bytes)
            result["has_heatmap"] = True
        except Exception:
            result["has_heatmap"] = False

    _write_json(_result_path(result["job_id"], "pairwise_results"), result)
    job_status_store[result["job_id"]] = result
    return result


@app.post("/pairwise-blast-alignment")
async def pairwise_blast_alignment(
    sequence: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    sequence_type: str = Form("dna"),
    target_type: str = Form("dna"),
    database: str = Form("nt"),
    alignment_mode: str = Form("global"),
    min_identity: float = Form(30.0),
    max_identity: float = Form(100.0),
    result_count: int = Form(50),
    expect_value: float = Form(10.0),
    megablast: bool = Form(True),
    email: Optional[str] = Form(None),
    max_total_sequences: int = Form(200),
):
    """Combined analysis: BLAST user sequences, fetch hits, then pairwise align all.

    Supports both file upload and text paste (FASTA format).

    Workflow:
    1. Parse user's multi-sequence FASTA (or pasted text)
    2. Run NCBI BLAST on each user sequence
    3. Fetch full sequences of BLAST hits from NCBI via Entrez
    4. Combine user sequences + BLAST hit sequences
    5. Run pairwise alignment on the combined set
    6. Generate: FASTA file, BLAST Excel, Identity Excel, Heatmap PNG, Filter log
    """
    _validate_common_params(
        sequence_type,
        target_type,
        database=database,
        min_identity=min_identity,
        max_identity=max_identity,
        result_count=result_count,
    )
    if alignment_mode not in {"global", "local"}:
        raise HTTPException(status_code=400, detail="alignment_mode must be global or local")
    if not (2 <= max_total_sequences <= 500):
        raise HTTPException(
            status_code=400, detail="max_total_sequences must be between 2 and 500"
        )
    if expect_value <= 0:
        raise HTTPException(status_code=400, detail="expect_value must be greater than 0")
    if not sequence and not file:
        raise HTTPException(status_code=400, detail="Either sequence text (FASTA) or file must be provided")

    input_info: Dict[str, Any] = {}
    if sequence:
        content_str = sequence.strip()
        input_info["source"] = "text"
    else:
        assert file is not None
        content = await file.read()
        try:
            content_str = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail="Uploaded file must be UTF-8 text") from exc
        safe_name = Path(file.filename or "combined.fasta").name
        upload_path = UPLOADS_DIR / f"{uuid.uuid4()}_{safe_name}"
        upload_path.write_bytes(content)
        input_info["source"] = "file"
        input_info["filename"] = safe_name
        input_info["saved_path"] = str(upload_path)

    job_id = str(uuid.uuid4())

    asyncio.create_task(
        run_combined_pairwise_analysis(
            job_id=job_id,
            content_str=content_str,
            sequence_type=sequence_type,
            target_type=target_type,
            database=database,
            alignment_mode=alignment_mode,
            min_identity=min_identity,
            max_identity=max_identity,
            result_count=result_count,
            expect_value=expect_value,
            megablast=megablast,
            email=email,
            max_total_sequences=max_total_sequences,
        )
    )

    return {
        "job_id": job_id,
        "status": "submitted",
        "message": "Combined BLAST + Pairwise analysis started",
        "program": choose_blast_program(sequence_type, target_type),
        "database": database,
        "input": input_info,
        "has_excel": False,
        "has_fasta": False,
        "has_blast_excel": False,
        "has_heatmap": False,
        "has_filter_log": False,
    }


@app.get("/results/{job_id}/excel")
async def download_result_excel(job_id: str):
    """Download the identity matrix as an Excel (.xlsx) file with timestamped filename."""
    if not HAS_OPENPYXL:
        raise HTTPException(
            status_code=501,
            detail="Excel export is not available. Install openpyxl: pip install openpyxl",
        )

    result = _load_job_result(job_id)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Try combined paths first, then pairwise, then blast
    for suffix in ("combined", "pairwise_blast", "pairwise", "blast"):
        excel_path = _result_path(job_id, suffix)
        excel_path = excel_path.with_suffix(".xlsx")
        if excel_path.exists():
            analysis_type = result.get("analysis_type", "analysis")
            db = result.get("database", "")
            db_suffix = f"_{db}" if db else ""
            filename = f"{analysis_type}{db_suffix}_{ts}.xlsx"
            return StreamingResponse(
                BytesIO(excel_path.read_bytes()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

    raise HTTPException(status_code=404, detail="Excel file not found for this job")


@app.get("/results/{job_id}/blast-excel")
async def download_blast_excel(job_id: str):
    """Download the BLAST results as an Excel (.xlsx) file."""
    if not HAS_OPENPYXL:
        raise HTTPException(
            status_code=501,
            detail="Excel export is not available. Install openpyxl: pip install openpyxl",
        )

    result = _load_job_result(job_id)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    for suffix in ("combined_blast", "blast"):
        excel_path = _result_path(job_id, suffix)
        excel_path = excel_path.with_suffix(".xlsx")
        if excel_path.exists():
            db = result.get("database", "")
            db_suffix = f"_{db}" if db else ""
            filename = f"blast_results{db_suffix}_{ts}.xlsx"
            return StreamingResponse(
                BytesIO(excel_path.read_bytes()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

    raise HTTPException(status_code=404, detail="BLAST Excel file not found for this job")


@app.get("/results/{job_id}/fasta")
async def download_result_fasta(job_id: str):
    """Download all sequences as a FASTA (.fasta) file."""
    result = _load_job_result(job_id)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    fasta_path = _result_path(job_id, "combined")
    fasta_path = fasta_path.with_suffix(".fasta")
    if fasta_path.exists():
        filename = f"combined_sequences_{ts}.fasta"
        return StreamingResponse(
            BytesIO(fasta_path.read_bytes()),
            media_type="text/plain",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    raise HTTPException(status_code=404, detail="FASTA file not found for this job")


@app.get("/results/{job_id}/heatmap")
async def download_result_heatmap(job_id: str):
    """Download the identity heatmap as a PNG image."""
    if not HAS_MATPLOTLIB:
        raise HTTPException(
            status_code=501,
            detail="Heatmap generation is not available. Install matplotlib: pip install matplotlib",
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Check both combined and pairwise heatmap paths
    for suffix in ("combined", "pairwise"):
        heatmap_path = _result_path(job_id, suffix)
        heatmap_path = heatmap_path.with_suffix(".png")
        if heatmap_path.exists():
            filename = f"identity_heatmap_{ts}.png"
            return StreamingResponse(
                BytesIO(heatmap_path.read_bytes()),
                media_type="image/png",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

    raise HTTPException(status_code=404, detail="Heatmap image not found for this job")


@app.get("/results/{job_id}/filter-log")
async def download_filter_log(job_id: str):
    """Download the filtered sequences log as a text file."""
    filter_path = _result_path(job_id, "combined_filter")
    filter_path = filter_path.with_suffix(".txt")
    if filter_path.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"filtered_sequences_{ts}.txt"
        return StreamingResponse(
            StringIO(filter_path.read_text()),
            media_type="text/plain",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    raise HTTPException(status_code=404, detail="Filter log not found for this job")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
